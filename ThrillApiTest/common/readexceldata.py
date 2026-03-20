'''
  @Project ：JungleApiTest 
  @File ：readexceldata.py 
  @Author ：lin20 
  @Date ：2025/9/19 11:43 
  @Describe：读取excel的测试数据生成测试参数
'''
from openpyxl import load_workbook

class cases:
    def __init__(self,id=None,category=None,title=None,method=None,url=None,data=None,expect=None,gpsinfo=None):
        self.id=id
        self.category=category
        self.title=title
        self.method=method
        self.url=url
        self.data=data
        self.expect=expect
        self.gpsinfo=gpsinfo



class doExcel:
    def __init__(self,file_name):

        self.file_name=file_name
        self.file=load_workbook(self.file_name)

        # self.sheet=self.file[sheet]
    def base_data(self,sheet,row_start=None,row_end=None):
        sheet_1=self.file[sheet]
        datas=[]
        try:

            if row_start is None:
                row_start = 2

            if row_end is None:
                    row_end = sheet_1.max_row

            for i in range(row_start,int(row_end)+1):
                case = cases()
                case.id=sheet_1.cell(row=i,column=1).value
                case.category=sheet_1.cell(row=i,column=2).value
                case.title=sheet_1.cell(row=i,column=3).value
                case.method=sheet_1.cell(row=i,column=4).value
                case.url=sheet_1.cell(row=i, column=5).value
                case.data=sheet_1.cell(row=i, column=6).value
                case.expect=sheet_1.cell(row=i, column=7).value
                datas.append(case)
        except Exception as e:
            print("读取数据报错：",e)
            raise e

        return datas

    def write_basedata(self,sheet,w_row,actual,result):
        sheet_1=self.file[sheet]
        sheet_1.cell(row=w_row,column=8).value=actual
        sheet_1.cell(row=w_row, column=9).value = result
        # self.sheet.cell(row,column).value = value_1

        self.file.save(filename=self.file_name)

    def charge_data(self,sheet,row_start=None,row_end=None):

        sheet_1=self.file[sheet]
        datas=[]
        try:
            if row_start is None:
                row_start = 2

            if row_end is None:
                row_end = sheet_1.max_row

            for i in range(row_start, int(row_end) + 1):
                case = cases()
                case.id=sheet_1.cell(row=i,column=1).value
                case.category=sheet_1.cell(row=i,column=2).value
                case.title=sheet_1.cell(row=i,column=3).value
                case.method=sheet_1.cell(row=i,column=4).value
                case.url=sheet_1.cell(row=i, column=5).value
                # normalize stored field name to `data` for consistency with other modules
                case.data = sheet_1.cell(row=i, column=6).value
                case.gpsinfo = sheet_1.cell(row=i, column=7).value
                case.expect=sheet_1.cell(row=i, column=8).value
                datas.append(case)
        except Exception as e:
            print("读取数据报错：", e)
            raise e

        return datas

    def write_chargedata(self,sheet,w_row,actual,result):
        sheet_1=self.file[sheet]
        sheet_1.cell(row=w_row,column=9).value=actual
        sheet_1.cell(row=w_row, column=10).value = result
        # self.sheet.cell(row,column).value = value_1

        try:
            self.file.save(filename=self.file_name)
        except PermissionError as e:
            # 常见于文件被其他程序占用（Excel 打开），记录并继续，不使测试因此中断
            print(f"无法写入 Excel 文件 {self.file_name}（可能被占用）：{e}")
        except Exception as e:
            # 其他写入错误仍然抛出以便上层知道
            print(f"保存 Excel 时发生异常: {e}")
            raise


if __name__ == '__main__':


    import route
    import generaterequestdata

    from common import readconfigfile
    from common import route
    import json

    rconf = readconfigfile.ReadConfig(route.conffile)
    read = doExcel(route.datafile)
    cases_2 = read.charge_data('charge',23,28)
    common_data = eval(rconf.get_value('data', 'common_data'))


    # req_data = eval(rconf.get_value('data', 'guest_data'))
    # guest_data = common_data
    # guest_data.update(req_data)
    # guest_url = "/user/guest"
    # req_guest = generaterequestdata.GenerateRequestData(url=guest_url, data=guest_data, method="POST")
    # resp = req_guest.reqapi()
    # result_guest = resp.json()
    # token=result_guest['data']['user']['authorization']
    # print(token)





    token="300456_nWhZrRYqnBvGbWicfhOnPi7LDo1ipHJRjJbHUxFeMiwmsIGMwW2WMzt5fVLo7y23MdRlbKXJd9BDmYGFGPrRaP3I0jIY19ILxTTi"

    for i in cases_2:



        querydata = common_data.copy()
        data=json.loads(i.data) if isinstance(i.data, str) else i.data
        querydata.update(data)
        s=generaterequestdata.GenerateRequestData(data=querydata, url=i.url, method=i.method, authorize=token)

        resp=s.reqapi()


        read.write_chargedata(sheet='charge', w_row=i.id + 1, actual=str(resp.json()),result='pass')
